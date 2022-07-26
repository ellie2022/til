# excel automation
# pip install openpyxl
import os.path

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def recalc_price(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        correted_price = cell.value * 0.9
        correted_price_cell = sheet.cell(row, 4)
        correted_price_cell.value = correted_price
        print(correted_price)
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    sep_filename = ['', '']
    sep_filename = os.path.splitext(filename)
    print(sep_filename)
    new_filename = sep_filename[0] + '2' + sep_filename[1]
    wb.save(new_filename)


recalc_price('transactions.xlsx')